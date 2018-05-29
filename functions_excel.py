# -*- coding: utf-8 -*-
"""
Created on Sat Feb 10 13:36:21 2018

@author: Zdenek
"""

import openpyxl 
import numpy as np

class Excel:
    def __Init__(self):
        self.worksheet = None
        self.workbook  = None
        self.filename  = None

# =============================================================================
# Creates new workbook with named worksheet
# =============================================================================
    def create_workbook_with_named_worksheet(self, workbook_name, worksheet_name):  
        self.filename = workbook_name          
        try:
            self.workbook  = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = worksheet_name            
        except Exception:
            print("Failed at function: create_workbook_with_named_worksheet()") 
            return None
        return self.workbook, self.worksheet
    
# =============================================================================
# Creates new workbook
# =============================================================================
    def create_workbook(self, workbook_name): 
        self.filename = workbook_name             
        try:
            self.workbook  = openpyxl.Workbook()
            self.worksheet = self.workbook.active          
        except Exception:
            print("Failed at function: create_workbook()") 
            return None
        return self.workbook, self.worksheet 

# =============================================================================
# Adds new worksheet in current workbook
# =============================================================================
    def add_worksheet_to_workbook(self, worksheet_name):  
        ll = self.workbook.get_sheet_names()       
        for name in ll:
            if name == worksheet_name:
                print("Failed at function: add_worksheet_to_workbook()") 
                print("Worksheet : {} already exists".format(worksheet_name)) 
                return None       
        self.worksheet = self.workbook.create_sheet(worksheet_name)             
        return self.workbook, self.worksheet
       
# =============================================================================
# Opens workbook and its nth worksheet by value
# =============================================================================
    def open_workbook(self, filename, nth_worksheet):
        self.filename = filename
        try:
            self.workbook  = openpyxl.load_workbook(filename) 
        except FileNotFoundError:
            print("Failed at function: open_workbook()") 
            print("File not found")  
            return None
        try:
            ll = self.workbook.get_sheet_names() 
            self.worksheet = self.workbook.get_sheet_by_name(ll[nth_worksheet])        
        except Exception:
            print("Failed at function: open_workbook()")   
            print("Worksheet_name number : {} does not exist".format(nth_worksheet))
            return None
        return self.workbook, self.worksheet

# =============================================================================
# Opens workbook and its worksheet by name
# =============================================================================
    def open_workbook_with_worksheet_by_name(self, filename, worksheet_name):
        self.filename = filename
        try:
            self.workbook  = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            print("Failed at function: open_workbook_with_worksheet_by_name()") 
            print("File not found") 
            return None
        try:
            self.worksheet = self.workbook.get_sheet_by_name(worksheet_name)       
        except Exception:
            print("Failed at function: open_workbook_with_worksheet_by_name()")   
            print("Worksheet with name : {} does not exist".format(worksheet_name))
            return None
        return self.workbook, self.worksheet  
    
# =============================================================================
# Opens worksheet by name in current workbook
# =============================================================================
    def open_worksheet_by_name(self, worksheet_name):       
        try:
            self.worksheet = self.workbook.get_sheet_by_name(worksheet_name)        
        except Exception:
            print("Failed at function: open_worksheet_by_name()")   
            print("Eorkbook with name : {} does not exist".format(worksheet_name))
            return None
        return self.workbook, self.worksheet 
    
# =============================================================================
# Opens worksheet by value in current workbook
# =============================================================================
    def open_worksheet_by_value(self, nth_worksheet):     
        try:
            ll = self.workbook.get_sheet_names()
            self.worksheet = self.workbook.get_sheet_by_name(ll[nth_worksheet])     
        except Exception:
            print("Failed at function: open_worksheet_by_value()")   
            print("Worksheet_name number : {} does not exist".format(nth_worksheet))
            return None   
        return self.workbook, self.worksheet 
    
# =============================================================================
# Writes to active worksheet's selected cell and save/or_not workbook   
# =============================================================================
    def write_to_active_worksheet(self, row, column, value, save_workbook = False):                
            self.worksheet.cell(column = column, row = row , value = value)
            if save_workbook:        
                self.workbook.save(self.filename)

# =============================================================================
# Writes range to active worksheet; cell start is starting point for range of values; and save/or_not  
# =============================================================================
    def write_range_to_active_worksheet(self, row_start, column, list_of_values, save_workbook = False):                          
        for position in np.arange(row_start, len(list_of_values)+row_start):            
            self.worksheet.cell(column = column, row = position, value = list_of_values[position - row_start])        
        if save_workbook:        
            self.workbook.save(self.filename)
        return position+1
# =============================================================================
# Reads data from active workseet"s cell
# =============================================================================
    def read_value_from_active_sheet(self, row, column): 
        cell = self.worksheet.cell(row=row, column=column)
        return cell.value
        
# =============================================================================
# Reads data from active workseet's range of cells; row_start and row_end define range    
# =============================================================================
    def read_row_range_from_active_sheet(self, row_start, row_end, column):
        _list = []
        for row in np.arange(row_start, row_end+1):
            _list.append(self.read_value_from_active_sheet(row, column))
        return _list
        
# =============================================================================
# # Saves active workbook 
# =============================================================================
    def save_active_workbook(self, filename=""):     
        if filename == "":           
            self.workbook.save(self.filename)
        else:
            self.workbook.save(filename)
            






















